from pathlib import Path
import json
import re
import time
import urllib.parse
import urllib.request
from html import unescape
from concurrent.futures import ThreadPoolExecutor, as_completed

from openpyxl import load_workbook
from openpyxl.comments import Comment


SOURCE = Path(r"C:\Users\maksim\Downloads\Soulstone Survivor Megasheet.xlsx")
OUTPUT = Path(r"C:\Users\maksim\SoulstoneMobilePrototype\Docs\Soulstone Survivor Megasheet RU.xlsx")
CACHE = Path(r"C:\Users\maksim\SoulstoneMobilePrototype\Docs\soulstone_megasheet_translation_cache.json")
ENABLE_MACHINE_TRANSLATION = True


EXACT = {
    "Primary Types": "Основные типы",
    "Character": "Персонаж",
    "Extra Types": "Дополнительные типы",
    "Skill name": "Название навыка",
    "Unlocked by": "Открывается через",
    "2nd Type": "Второй тип",
    "Extra Types": "Доп. типы",
    "Debuffs": "Дебаффы",
    "Buffs": "Баффы",
    "Trait": "Черта",
    "Titan Power": "Сила титана",
    "Description": "Описание",
    "Requirements": "Требования",
    "Details": "Детали",
    "Effect": "Эффект",
    "Damage": "Урон",
    "Stacks": "Стаки",
    "Chance": "Шанс",
    "Average damage": "Средний урон",
    "Average stacks": "Средние стаки",
    "Formulas:": "Формулы:",
    "Health": "Здоровье",
    "Armor": "Броня",
    "Body mass": "Масса тела",
    "Armor Penetration": "Пробивание брони",
    "Attack speed": "Скорость атаки",
    "Movement": "Скорость движения",
    "Multiplier": "Множитель",
    "Expected value": "Ожидаемое значение",
    "Guaranteed": "Гарантировано",
    "Quality Standards": "Стандарты качества",
    "Weight": "Вес",
    "Chance of best": "Шанс лучшего",
    "Rarity": "Редкость",
    "Common": "Обычная",
    "Uncommon": "Необычная",
    "Rare": "Редкая",
    "Epic": "Эпическая",
    "Legendary": "Легендарная",
    "Default": "По умолчанию",
    "Both": "Оба",
    "Commoner": "Простолюдин",
    "Arcane": "Тайная магия",
    "Blast": "Взрыв",
    "Bomb": "Бомба",
    "Chaos": "Хаос",
    "Earth": "Земля",
    "Electric": "Электричество",
    "Fire": "Огонь",
    "Holy": "Свет",
    "Ice": "Лёд",
    "Nature": "Природа",
    "Projectile": "Снаряд",
    "Shadow": "Тьма",
    "Slam": "Удар",
    "Swing": "Взмах",
    "Thrust": "Укол",
    "Aura": "Аура",
    "Chain": "Цепь",
    "Summon": "Призыв",
    "Area": "Область",
    "Burst": "Всплеск",
    "Empowering": "Усиление",
    "Frontal": "Фронтальная",
    "Lasting": "Длительная",
    "Missile": "Снаряд/ракета",
    "Static": "Статичная",
    "Secret": "Секретная",
    "Magical": "Магическая",
    "Physical": "Физическая",
    "Arcane Weaver": "Ткач тайной магии",
    "Houndmaster": "Псарь",
    "Assassin": "Ассасин",
    "Chaoswalker": "Странник хаоса",
    "Monkey King": "Царь обезьян",
    "Elementalist": "Элементалист",
    "Pyromancer": "Пиромант",
    "Paladin": "Паладин",
    "Death Knight": "Рыцарь смерти",
    "Beastmaster": "Повелитель зверей",
    "Sentinel": "Страж",
    "Necromancer": "Некромант",
    "Barbarian": "Варвар",
    "Spellblade": "Магический клинок",
    "Legionnaire": "Легионер",
    "Engineer": "Инженер",
    "Myrmidon": "Мирмидонец",
    "Demon Hunter": "Охотник на демонов",
    "Cursed Captain": "Проклятый капитан",
    "Shaman": "Шаман",
    "Machinist": "Механик",
    "Samurai": "Самурай",
    "Blacksmith": "Кузнец",
    "Bleed": "Кровотечение",
    "Burn": "Горение",
    "Poison": "Яд",
    "Doom": "Рок",
    "Slow": "Замедление",
    "Dazed": "Ошеломление",
    "Brittle": "Хрупкость",
    "Exposed": "Раскрытие",
    "Fragility": "Уязвимость",
    "Weakness": "Слабость",
    "Shattered": "Раскол",
    "Disoriented": "Дезориентация",
    "Disarray": "Разлад",
    "Distracted": "Отвлечение",
    "Debilitated": "Ослабление",
    "Crippled": "Увечье",
    "Finesse": "Изящество",
    "Bulwark": "Оплот",
    "Aptitude": "Способность",
    "Prowess": "Мастерство",
    "Resilience": "Стойкость",
    "Form": "Форма",
    "Haste": "Спешка",
    "Colossal": "Колоссальность",
    "Vigor": "Живучесть",
    "Piercing": "Пробивание",
    "Brutal": "Жестокость",
    "Precise": "Точность",
    "Impactful": "Мощное попадание",
    "Devastating": "Сокрушение",
    "Chaotic": "Хаотичность",
    "Cunning": "Хитрость",
    "Ancestry": "Родословная",
    "Ambush": "Засада",
    "Radiance": "Сияние",
    "Electrified": "Электризация",
    "Icy Veins": "Ледяные жилы",
    "Purity": "Чистота",
    "Ammunition": "Боеприпасы",
    "Retaliation": "Возмездие",
    "Cursed": "Проклятость",
    "Fortitude": "Крепость",
    "Gunpowder": "Порох",
    "Condemned": "Осуждение",
    "Divinity": "Божественность",
    "Overlord": "Оверлорд",
    "Normal foes": "Обычные враги",
    "Passive powers": "Пассивные силы",
    "Artifact Powers": "Силы артефактов",
    "Divine legacy": "Божественное наследие",
    "Trivia": "Примечание",
    "Legendary choices": "Легендарные выборы",
    "Value": "Значение",
    "on hit": "при попадании",
    "on crit": "при крите",
    "on disarray": "при разладе",
    "Severe Cold": "Сильный холод",
    "Arachnid Frenzy": "Паучье бешенство",
    "Frenzied Bites": "Бешеные укусы",
    "Potent Toxin": "Сильный токсин",
    "Web of Confusion": "Паутина замешательства",
    "Fangs of Doom": "Клыки рока",
    "Spider Senses": "Паучье чутьё",
    "Ultimate Assassin": "Идеальный ассасин",
    "Synchrony": "Синхрония",
    "Unconv Start": "Необычный старт",
    "Lord's Bane": "Погибель лорда",
    "Last Resort": "Последний шанс",
    "Headstart": "Фора",
    "Head Start": "Фора",
    "All or Nothing": "Всё или ничего",
    "Impaler": "Пронзатель",
    "Surefooted": "Уверенная поступь",
    "Executioner": "Палач",
    "Dash Mastery": "Мастерство рывка",
    "Center of Attention": "Центр внимания",
    "Commanding Presence": "Властное присутствие",
    "Misforune Embrace": "Объятия несчастья",
    "Misfortune Embrace": "Объятия несчастья",
    "Unwavering Persistence": "Непоколебимое упорство",
    "Quick Reload": "Быстрая перезарядка",
}


PHRASES = [
    ("Y stacks of Bleed, X/10 per tick over 10s", "Y стаков кровотечения, X/10 за тик в течение 10 сек"),
    ("Y stacks of Burn, X/8 per tick over 4s", "Y стаков горения, X/8 за тик в течение 4 сек"),
    ("Y stacks of Poison, X/15 per tick over 15s, spread on death", "Y стаков яда, X/15 за тик в течение 15 сек, распространяется при смерти"),
    ("Y stacks of Doom, X damage, X/4 aoe damage on death/proc", "Y стаков рока, X урона, X/4 AoE урона при смерти/срабатывании"),
    ("Decrease movement speed by X * 1%", "Снижает скорость передвижения на X * 1%"),
    ("Increase crit chance by X * 1%", "Повышает шанс крита на X * 1%"),
    ("Increase base damage by X * 1", "Повышает базовый урон на X * 1"),
    ("Increase base damage by up to X * 4%", "Повышает базовый урон до X * 4%"),
    ("Increase damage taken by X * 1%", "Повышает получаемый урон на X * 1%"),
    ("Increase dot and doom damage taken by X * 1.2%", "Повышает получаемый DoT-урон и урон рока на X * 1.2%"),
    ("Decrease armor by X * 1", "Снижает броню на X * 1"),
    ("Increase crit damage taken by X * 1%", "Повышает получаемый критический урон на X * 1%"),
    ("Increase miss chance by X * 0.5%", "Повышает шанс промаха на X * 0.5%"),
    ("Decrease damage by X * 1%", "Снижает урон на X * 1%"),
    ("Decrease area by X * 0.5%", "Снижает область на X * 0.5%"),
    ("Multicast + X * 1%", "Мультикаст + X * 1%"),
    ("Block + X", "Блок + X"),
    ("Crit damage + X * 2%", "Критический урон + X * 2%"),
    ("Damage + X * 1%", "Урон + X * 1%"),
    ("Armor + X", "Броня + X"),
    ("Crit chance + X * 1%", "Шанс крита + X * 1%"),
    ("Movement + X * 1%", "Скорость движения + X * 1%"),
    ("Area + X * 1%", "Область + X * 1%"),
    ("HP + X * 0.5", "HP + X * 0.5"),
    ("Increase damage by 20% per enemy hit", "Увеличивает урон на 20% за каждого поражённого врага"),
    ("Increase damage by up to 50% for nearby enemies", "Увеличивает урон до 50% по ближайшим врагам"),
    ("Increase damage by up to 100% for far enemies", "Увеличивает урон до 100% по дальним врагам"),
    ("Increase damage by up to 50% for enemies close to the center of the aoe", "Увеличивает урон до 50% по врагам рядом с центром AoE"),
    ("Deals 3x damage on crit hit", "Наносит 3x урон при критическом попадании"),
    ("Base monster armor is 100", "Базовая броня монстра равна 100"),
    ("Damage formula: 2/((Armor/100) + 1). Each armor point above 100 increases effective health by 0.5%", "Формула урона: 2/((Броня/100)+1). Каждая единица брони выше 100 повышает эффективное здоровье на 0.5%"),
    ("for each armor point under 100 the monster will take 1% more damage", "за каждую единицу брони ниже 100 монстр получает на 1% больше урона"),
    ("The cast speed cap is 0.15s (displayed as 0.2s)", "Кап скорости применения — 0.15 сек, отображается как 0.2 сек"),
    ("Rarity is rolled first and then possible cards within the rarity is rolled", "Сначала бросается редкость, затем выбираются возможные карты внутри этой редкости"),
]


WORD_REPLACEMENTS = {
    "Skills": "Навыки",
    "Skill": "Навык",
    "Debuff": "Дебафф",
    "Debuffs": "Дебаффы",
    "Buff": "Бафф",
    "Buffs": "Баффы",
    "Summons": "Призывы",
    "summons": "призывы",
    "Enemies": "Враги",
    "enemies": "враги",
    "enemy": "враг",
    "Boss": "Босс",
    "boss": "босс",
    "Elites": "Элита",
    "Lords": "Лорды",
    "Damage": "Урон",
    "damage": "урон",
    "HP": "HP",
    "effective": "эффективное",
    "Armor": "Броня",
    "AoE": "область",
    "RepeatChance": "шанс повторного применения",
    "DoubleAttackChance": "шанс двойной атаки",
    "DoT": "периодический урон",
    "Burn": "Горение",
    "Bleed": "Кровотечение",
    "Doom": "Рок",
    "Endless": "Бесконечный режим",
    "Warlords": "Полководцы",
    "Void": "Пустоты",
    "EXP": "Опыт",
    "exp": "опыт",
    "level": "уровень",
    "Spawn": "Спавн",
    "Gap": "Интервал",
    "Oceanic": "Океанический",
    "Wisdom": "Мудрость",
    "Imminent": "Неминуемая",
    "Material": "Материалы",
    "Harvest": "Сбор",
    "Singular": "Единый",
    "Focus": "Фокус",
    "build": "билд",
    "Quick": "Быстрый",
    "Reload": "Перезарядка",
    "Sub": "До",
    "Cycle": "Цикл",
    "Min-Max": "Мин-макс",
    "situational": "ситуативно",
    "weghts": "веса",
    "weights": "веса",
    "shifted": "смещены",
    "interesting": "интересно",
    "trap": "ловушка",
    "skills": "навыки",
    "active": "активные",
    "locks": "блокировки",
    "nice": "хороший",
    "access": "доступ",
    "death": "смерть",
    "protection": "защита",
    "useful": "полезно",
    "defensive": "защитные",
    "options": "варианты",
    "Demon": "Демон",
    "hunter": "охотник",
    "Hunter": "Охотник",
    "runic": "руническая",
    "power": "сила",
    "Cooldown": "Кулдаун",
    "cooldown": "кулдаун",
    "CD": "КД",
    "cd": "кд",
    "Cast": "Каст",
    "cast": "каст",
    "Crit": "Крит",
    "crit": "крит",
    "chance": "шанс",
    "Chance": "Шанс",
    "Movement": "Движение",
    "movement": "движение",
    "Speed": "Скорость",
    "speed": "скорость",
    "Area": "Область",
    "area": "область",
    "Duration": "Длительность",
    "duration": "длительность",
    "Stacks": "Стаки",
    "stacks": "стаки",
    "stack": "стак",
    "Source": "Источник",
    "source": "источник",
    "Requirement": "Требование",
    "Requires": "Требует",
    "requires": "требует",
    "Apply": "Наложить",
    "apply": "наложить",
    "multiplier": "множитель",
    "Multiplier": "Множитель",
    "Random": "Случайно",
    "random": "случайно",
    "Legendary": "Легендарная",
    "Epic": "Эпическая",
    "Rare": "Редкая",
    "Uncommon": "Необычная",
    "Common": "Обычная",
    "Damage reduction": "Снижение урона",
    "Rain": "Дождь",
    "Arrows": "Стрел",
    "Arrow": "Стрела",
    "Dreadful": "Ужасная",
    "Grenade": "Граната",
    "Eviscerate": "Потрошение",
    "Shrapnel": "Осколочная",
    "QT": "Кью-Ти",
    "Bombuddy": "Бомбарик",
    "Explosion": "Взрыв",
    "Quicksand": "Зыбучие пески",
    "Storm": "Буря",
    "Spears": "Копий",
    "Spear": "Копьё",
    "Celestial": "Небесное",
    "Retribution": "Возмездие",
    "Battle": "Боевой",
    "Shout": "Крик",
    "Radiant": "Сияющие",
    "Clones": "Клоны",
    "Uppercut": "Апперкот",
    "Subdue": "Подчинение",
    "Slice": "Рассечь",
    "Dice": "Нарезать",
    "Ancestral": "Родовое",
    "Empowerment": "Усиление",
    "Assault": "Натиск",
    "Beam": "Луч",
    "Blade": "Клинок",
    "Blades": "Клинки",
    "Conjuration": "Сотворение",
    "Disc": "Диск",
    "Ammo": "Боеприпасы",
    "Box": "Ящик",
    "Overload": "Перегрузка",
    "Buckshot": "Картечь",
    "Explosive": "Взрывной",
    "Shot": "Выстрел",
    "Sentry": "Турель",
    "Mechanical": "Механический",
    "Spider": "Паук",
    "Bear": "Медвежий",
    "Trap": "Капкан",
    "Bombardment": "Бомбардировка",
    "Barrage": "Залп",
    "Canon": "Пушечный",
    "Cannon": "Пушечный",
    "Cluster": "Кластерная",
    "Debilitating": "Ослабляющая",
    "Bloodbane": "Кровавая погибель",
    "Crossbow": "Арбалет",
    "Bloodlust": "Жажда крови",
    "Carrion": "Падальные",
    "Crows": "Вороны",
    "Onslaught": "Натиск",
    "Bolt": "Снаряд",
    "Eruption": "Извержение",
    "Abyssal": "Бездонный",
    "Fissure": "Разлом",
    "Demolish": "Снос",
    "Drilling": "Буровая",
    "Rig": "Установка",
    "Shield": "Щит",
    "Surge": "Всплеск",
    "Lightning": "Молния",
    "Call": "Призыв",
    "Strike": "Удар",
    "Energizer": "Зарядник",
    "God": "Божий",
    "Rebuke": "Укор",
    "Blazehound": "Пламенная гончая",
    "Blazing": "Пылающий",
    "Ray": "Луч",
    "Combustion": "Сгорание",
    "Conflagrate": "Пожар",
    "Ember": "Угольный",
    "Discharge": "Разряд",
    "Fiery": "Огненные",
    "Beacon": "Маяк",
    "Light": "Света",
    "Bull": "Бычий",
    "Charge": "Рывок",
    "Cleansing": "Очищающий",
    "Hammer": "Молот",
    "Justice": "Правосудия",
    "Arctic": "Арктический",
    "Avalanche": "Лавина",
    "Blizzard": "Метель",
    "Bone": "Костяной",
    "Pillar": "Столб",
    "Freezing": "Ледяной",
    "Blow": "Удар",
    "Acid": "Кислотный",
    "Moose": "Лось",
    "Bestial": "Звериная",
    "Wrath": "Ярость",
    "Cave": "Пещерный",
    "Cobra": "Кобра",
    "Totem": "Тотем",
    "Contagion": "Заражение",
    "Devourer": "Пожиратель",
    "Bladed": "Клинковый",
    "Chakram": "Чакрам",
    "Build": "Построить",
    "Ballista": "Баллиста",
    "Lead": "Свинец",
    "Camor": "Камор",
    "Crushing": "Сокрушительная",
    "Darkness": "Тьма",
    "Rune": "Руна",
    "Dark": "Тёмный",
    "Swarm": "Рой",
    "Death": "Смерть",
    "Decay": "Тление",
    "Vortex": "Вихрь",
    "Gathering": "Сгущающаяся",
    "Anvil": "Наковальня",
    "Mace": "Булава",
    "Shatter": "Раскол",
    "Impale": "Пронзание",
    "Reaping": "Жатва",
    "Slash": "Разрез",
    "Backstab": "Удар в спину",
    "Barbaric": "Варварские",
    "Cleavers": "Тесаки",
    "Flesh": "Плоть",
    "Flurry": "Шквал",
    "Skeletal": "Скелетные",
    "Warriors": "Воины",
    "Poison": "Ядовитый",
    "Puddle": "Лужа",
    "Double": "Двойной",
    "Glacier": "Ледник",
    "Power": "Сила",
    "Powers": "Силы",
    "Adds": "Добавляет",
    "adds": "добавляет",
    "based": "основано",
    "distance": "расстояние",
    "nearby": "близкие",
    "far": "дальние",
    "near": "рядом",
    "linear": "линейная",
    "front": "перед",
    "heals": "лечит",
    "hits": "попадает",
    "static": "статичный",
    "tentacle": "щупальце",
    "totem": "тотем",
    "consumption": "расход",
    "consume": "расходует",
    "direct": "прямой",
    "without": "без",
    "reaches": "достигает",
    "soft": "мягкий",
    "hard": "жёсткий",
    "exceptions": "исключения",
    "Titans": "Титаны",
    "Event": "Событие",
    "bosses": "боссы",
    "sources": "источники",
    "without": "без",
    "affects": "влияет",
    "active": "активные",
    "consumable": "расходуемые",
    "multiplicative": "мультипликативный",
    "bonus": "бонус",
    "ticks": "тики",
    "same": "та же",
    "overall": "общий",
    "Elemental": "Стихийный",
    "Flow": "Поток",
}


COMMENTS = {
    "Glossary": ("A1", "Ключевой лист-словарь: типы навыков, статусы, синергии, редкости и капы. Используй как основной источник терминов."),
    "Enemy stat": ("A1", "Таблица оверлорд-скейлинга врагов. Для нашего прототипа использованы относительные ориентиры HP/урона/скорости, а не гигантские factorial-значения поздней игры."),
    "Enemy stat (9g3)": ("A1", "Более мягкая таблица статов врагов; полезнее для раннего прототипа и мобильного баланса."),
    "Armor": ("B1", "Главная формула брони: выше 100 броня повышает effective HP, ниже 100 увеличивает входящий урон."),
    "Cast frequency": ("A1", "Важный ориентир: скорость каста имеет нижний кап около 0.15 сек. В нашем проекте не стоит делать кулдауны ниже этого порога."),
    "Multicast": ("A1", "Мультикаст считается как ожидаемое число повторов. В нашем проекте ближайший аналог — RepeatChance/DoubleAttackChance."),
    "Debuff synergies": ("A1", "Справочник синергий статусов: шанс наложения, средний урон и стаки."),
    "Stat interaction": ("A1", "Показывает, какие статы влияют на прямой урон, DoT, призывы и пассивные эффекты."),
    "Skill Scaling": ("A1", "Навыки могут скейлиться от баффов/дебаффов. Это хороший референс для будущей системы синергий второго уровня."),
    "All skills": ("A1", "Общий список навыков и их тегов/эффектов. Использован как референс для типов атак в нашем проекте."),
    "Artifact Powers": ("A1", "Поздняя мета-система. Пока не внедрялась, но полезна как будущий ориентир."),
    "Titan Powers": ("A1", "Продвинутые модификаторы/силы. Пока не внедрялись; можно использовать позже для affix/run modifiers."),
}


def translate(value: str) -> str:
    if not isinstance(value, str) or value.startswith("="):
        return value

    if value in EXACT:
        return EXACT[value]

    result = value
    for source, target in PHRASES:
        result = result.replace(source, target)

    for source, target in sorted(EXACT.items(), key=lambda item: len(item[0]), reverse=True):
        result = result.replace(source, target)

    for source, target in sorted(WORD_REPLACEMENTS.items(), key=lambda item: len(item[0]), reverse=True):
        result = re.sub(rf"\b{re.escape(source)}\b", target, result)

    return result


LATIN_RE = re.compile(r"[A-Za-z]{3,}")


def has_latin(value: object) -> bool:
    return isinstance(value, str) and not value.startswith("=") and bool(LATIN_RE.search(value))


def should_machine_translate(value: str) -> bool:
    if not has_latin(value):
        return False

    stripped = value.strip()
    if not stripped:
        return False

    # Do not waste translation calls on formulas, roman numerals, coordinates, or mostly numeric labels.
    if re.fullmatch(r"[IVXLCM0-9# .()+*/%-]+", stripped):
        return False

    return True


def load_cache() -> dict[str, str]:
    if CACHE.exists():
        return json.loads(CACHE.read_text(encoding="utf-8"))
    return {}


def save_cache(cache: dict[str, str]) -> None:
    CACHE.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


def machine_translate(text: str, cache: dict[str, str]) -> str:
    if text in cache:
        return cache[text]

    # MyMemory is used only as a fallback for cells not covered by the local glossary.
    # Keep each request compact; very long design notes are translated line by line.
    if len(text) > 450 or "\n" in text:
        parts = text.split("\n")
        translated_parts = [machine_translate(part, cache) if part.strip() else part for part in parts]
        result = "\n".join(translated_parts)
        cache[text] = result
        return result

    try:
        url = "https://translate.googleapis.com/translate_a/single?client=gtx&sl=en&tl=ru&dt=t&q=" + urllib.parse.quote(text)
        with urllib.request.urlopen(url, timeout=12) as response:
            payload = json.loads(response.read().decode("utf-8", errors="replace"))
        result = "".join(part[0] for part in payload[0] if part and part[0]) or text
        result = unescape(result)
    except Exception:
        try:
            url = "https://api.mymemory.translated.net/get?q=" + urllib.parse.quote(text) + "&langpair=en|ru"
            with urllib.request.urlopen(url, timeout=12) as response:
                payload = json.loads(response.read().decode("utf-8", errors="replace"))
            result = payload.get("responseData", {}).get("translatedText") or text
            result = unescape(result)
        except Exception:
            result = text

    cache[text] = result
    time.sleep(0.01)
    return result


def remote_translate_uncached(text: str) -> str:
    if len(text) > 450 or "\n" in text:
        return "\n".join(remote_translate_uncached(part) if part.strip() else part for part in text.split("\n"))

    try:
        url = "https://translate.googleapis.com/translate_a/single?client=gtx&sl=en&tl=ru&dt=t&q=" + urllib.parse.quote(text)
        with urllib.request.urlopen(url, timeout=10) as response:
            payload = json.loads(response.read().decode("utf-8", errors="replace"))
        return unescape("".join(part[0] for part in payload[0] if part and part[0]) or text)
    except Exception:
        try:
            url = "https://api.mymemory.translated.net/get?q=" + urllib.parse.quote(text) + "&langpair=en|ru"
            with urllib.request.urlopen(url, timeout=10) as response:
                payload = json.loads(response.read().decode("utf-8", errors="replace"))
            return unescape(payload.get("responseData", {}).get("translatedText") or text)
        except Exception:
            return text


def translate_remaining_cells(wb, cache: dict[str, str]) -> None:
    values: set[str] = set()

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if should_machine_translate(cell.value):
                    values.add(cell.value)

    missing = [value for value in values if value not in cache]

    if missing:
        with ThreadPoolExecutor(max_workers=24) as executor:
            futures = {executor.submit(remote_translate_uncached, value): value for value in missing}

            completed = 0
            for future in as_completed(futures):
                source = futures[future]
                try:
                    cache[source] = future.result()
                except Exception:
                    cache[source] = source

                completed += 1
                if completed % 200 == 0:
                    save_cache(cache)

        save_cache(cache)

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if should_machine_translate(cell.value):
                    cell.value = cache.get(cell.value, cell.value)


def translate_workbook_text(wb, cache: dict[str, str]) -> None:
    if not ENABLE_MACHINE_TRANSLATION:
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and not cell.value.startswith("="):
                        cell.value = translate(cell.value)
        return

    machine_values: set[str] = set()

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value

                if not isinstance(value, str) or value.startswith("="):
                    continue

                if value in EXACT:
                    continue

                if should_machine_translate(value):
                    machine_values.add(value)

    missing = [value for value in machine_values if value not in cache]

    if missing:
        with ThreadPoolExecutor(max_workers=24) as executor:
            futures = {executor.submit(remote_translate_uncached, value): value for value in missing}

            completed = 0
            for future in as_completed(futures):
                source = futures[future]

                try:
                    cache[source] = future.result()
                except Exception:
                    cache[source] = source

                completed += 1
                if completed % 200 == 0:
                    save_cache(cache)

        save_cache(cache)

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value

                if not isinstance(value, str) or value.startswith("="):
                    continue

                if value in EXACT:
                    cell.value = EXACT[value]
                elif should_machine_translate(value):
                    cell.value = translate(cache.get(value, value))
                else:
                    cell.value = translate(value)


def main() -> None:
    wb = load_workbook(SOURCE)
    cache = load_cache()

    translate_workbook_text(wb, cache)

    for sheet_name, (cell_ref, text) in COMMENTS.items():
        if sheet_name in wb.sheetnames:
            wb[sheet_name][cell_ref].comment = Comment(text, "Codex")

    if "RU Summary" in wb.sheetnames:
        del wb["RU Summary"]

    if "Needs Review" in wb.sheetnames:
        del wb["Needs Review"]

    ws = wb.create_sheet("RU Summary", 0)
    rows = [
        ["Раздел", "Вывод для нашего проекта"],
        ["Статы врагов", "Используем мягкий ранний скейлинг: обычные враги низкое HP/урон, танки примерно x2-3 HP, элита x4-5 HP, мини-босс x10+ HP."],
        ["Броня", "Броня 100 = нейтрально; выше 100 повышает эффективное здоровье, ниже 100 повышает входящий урон."],
        ["Частота каста", "Не опускаем эффективный кулдаун ниже 0.15 сек. Быстрые навыки слабее за удар, тяжёлые атаки по области имеют выше урон и дольше КД."],
        ["Мультикаст", "В нашем проекте аналог — шанс повторного применения и шанс двойной атаки. Для быстрых навыков шанс ниже, для оружия ассасина/лучника можно выше."],
        ["Призывы", "Навыки призыва должны иметь лимит активных существ, здоровье, урон, скорость и агро врагов. Некромант настроен по этой логике."],
        ["Статусы", "Эффекты периодического урона: горение короткое, яд длинный, кровотечение среднее, рок срабатывает как накопленный урон."],
    ]

    for row_index, row in enumerate(rows, start=1):
        for column_index, value in enumerate(row, start=1):
            ws.cell(row_index, column_index).value = value

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 120
    ws["A1"].comment = Comment("Этот лист добавлен Codex: короткая русская выжимка по применению таблицы к нашему прототипу.", "Codex")

    review = wb.create_sheet("Needs Review", 1)
    review.append(["Лист", "Ячейка", "Текст, где ещё есть латиница"])
    for source_ws in wb.worksheets:
        if source_ws.title == "Needs Review":
            continue
        for row in source_ws.iter_rows():
            for cell in row:
                if has_latin(cell.value):
                    review.append([source_ws.title, cell.coordinate, cell.value])
    review.column_dimensions["A"].width = 26
    review.column_dimensions["B"].width = 12
    review.column_dimensions["C"].width = 120

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    save_cache(cache)
    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
